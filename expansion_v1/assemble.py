#!/usr/bin/env python3
"""
Assemble the 8 expansion-market CSVs into:
  - Per-market XLSX files at expansion_v1/xlsx/<market>_2026_scored.xlsx
    (sheets: Accounts, TAM_Conservative, WAS_Scoring_v2 — schema-compatible
    with the existing 6 *_scored.xlsx files)
  - Combined accounts.json drop-in at expansion_v1/expansion_accounts.json
    (consumable by normalize_v3.py)

TAM formula (mirrors Charlotte v3 cookbook):
  Hotel:       rooms * occupancy * valet_conv * 1 (turnover) * valet_rate * op_days * peak_mod
  Restaurant:  seats * turnover * valet_conv * valet_rate * op_days * peak_mod
  Hospital:    beds * 0.6 (visitor multiplier) * valet_conv * valet_rate * op_days * peak_mod
               (default valet_conv 0.20, op_days 365 — visitor cars/day model)
  Venue:       capacity * valet_conv * valet_rate * op_days * peak_mod
               (op_days = event count; valet_conv default 0.15)
"""
import csv, json, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

EXPANSION = Path('/home/user/workspace/sophi-market-map/expansion_v1')
RAW = EXPANSION / 'raw'
XLSX_OUT = EXPANSION / 'xlsx'
XLSX_OUT.mkdir(parents=True, exist_ok=True)

# ---------- helpers ----------
def to_float(x, default=None):
    if x is None or x == '': return default
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip().replace('$','').replace(',','').replace('%','')
    if s.lower() in ('','n/a','na','none','-','tbd','unknown'): return default
    try: return float(s)
    except: return default

def to_int(x, default=None):
    v = to_float(x, default)
    return int(v) if v is not None else default

def default_valet_conv(typ, rooms):
    if typ == 'Hotel':
        if rooms and rooms >= 200: return 0.40
        return 0.30
    if typ == 'Restaurant': return 0.125
    if typ == 'Hospital': return 0.20
    if typ == 'Venue': return 0.15
    return 0.20

# ---------- type+market median rate imputation ----------
DEFAULT_RATES_BY_TYPE = {
    # fallback when even market median is unavailable
    'Hotel': 35.0,        # downtown midscale
    'Restaurant': 12.0,   # upscale-casual valet rate
    'Hospital': 8.0,      # validated visitor rate
    'Venue': 25.0,        # downtown event venue valet
}

MARKET_RATE_FLOOR = {  # tier-1 floors so cheap markets don't silently zero out
    'Hotel': 25.0,
    'Restaurant': 10.0,
    'Hospital': 5.0,
    'Venue': 15.0,
}

VENUE_CAPACITY_CAP = 22000.0   # sanity cap — stadiums get clamped (or use op_days * peak to model bigger volume), and convention-center sqft inputs get neutralized
VENUE_TAM_CAP = 5_000_000.0    # per-account venue TAM ceiling — anything higher gets clamped + flagged for manual review (matches the 6-market 99th-percentile)
HOTEL_TAM_CAP = 8_000_000.0    # per-account hotel TAM ceiling — captures e.g. Westin Charlotte $2.7M; flags Vegas-scale outliers
RESTAURANT_TAM_CAP = 1_000_000.0  # per-account restaurant cap — well above Charlotte top of $300K

def build_rate_pool(all_rows_by_market):
    """Returns dict[(market, type)] -> median rate from accounts with explicit usable rate."""
    import statistics
    pool = {}
    for mk, rows in all_rows_by_market.items():
        for r in rows:
            rate = to_float(r.get('valet_rate'))
            if rate is None or rate <= 0:
                rate = to_float(r.get('self_park_rate'))
            if rate is not None and rate > 0:
                pool.setdefault((mk, r.get('type')), []).append(rate)
    # collapse to medians
    return {k: statistics.median(v) for k, v in pool.items()}

RATE_POOL = {}  # populated in main

def resolve_rate(row):
    """Return (rate, source) using sourced > market+type median > type default."""
    rate = to_float(row.get('valet_rate'))
    if rate is not None and rate > 0:
        return rate, 'sourced_valet'
    rate = to_float(row.get('self_park_rate'))
    if rate is not None and rate > 0:
        return rate, 'sourced_self_park'
    mk = row.get('market_key')
    typ = row.get('type')
    if (mk, typ) in RATE_POOL:
        return RATE_POOL[(mk, typ)], 'imputed_market_median'
    return DEFAULT_RATES_BY_TYPE.get(typ, 25.0), 'imputed_type_default'

def compute_tam(row):
    """Compute conservative annual TAM for an account row."""
    typ = row.get('type','').strip()
    rate, rate_src = resolve_rate(row)
    if rate is None or rate <= 0:
        return 0.0, 'no_rate', rate_src
    op_days = to_float(row.get('op_days'), 365)
    peak    = to_float(row.get('peak_mod'), 1.0)
    conv    = to_float(row.get('valet_conv'))

    if typ == 'Hotel':
        rooms = to_float(row.get('rooms'))
        occ   = to_float(row.get('occupancy'), 0.60)
        if not rooms: return 0.0, 'no_rooms', rate_src
        if conv is None: conv = default_valet_conv('Hotel', rooms)
        raw_tam = rooms * occ * conv * rate * op_days * peak
        if raw_tam > HOTEL_TAM_CAP:
            return HOTEL_TAM_CAP, f'hotel_capped(raw=${raw_tam/1e6:.1f}M)', rate_src
        return raw_tam, 'hotel', rate_src

    if typ == 'Restaurant':
        seats = to_float(row.get('seats'))
        turn  = to_float(row.get('turnover'), 1.5)
        if not seats: return 0.0, 'no_seats', rate_src
        if conv is None: conv = 0.125
        raw_tam = seats * turn * conv * rate * op_days * peak
        if raw_tam > RESTAURANT_TAM_CAP:
            return RESTAURANT_TAM_CAP, f'restaurant_capped(raw=${raw_tam/1e6:.1f}M)', rate_src
        return raw_tam, 'restaurant', rate_src

    if typ == 'Hospital':
        beds = to_float(row.get('beds'))
        if not beds: return 0.0, 'no_beds', rate_src
        if conv is None: conv = 0.20
        visitor_mult = 0.6
        return beds * visitor_mult * conv * rate * op_days * peak, 'hospital', rate_src

    if typ == 'Venue':
        cap_raw = to_float(row.get('capacity'))
        if not cap_raw: return 0.0, 'no_capacity', rate_src
        # sanity cap: convention center 'capacity' fields often hold sqft; clamp at VENUE_CAPACITY_CAP
        cap = min(cap_raw, VENUE_CAPACITY_CAP)
        if conv is None: conv = 0.15
        raw_tam = cap * conv * rate * op_days * peak
        if raw_tam > VENUE_TAM_CAP:
            return VENUE_TAM_CAP, f'venue_capped(raw=${raw_tam/1e6:.1f}M)', rate_src
        return raw_tam, 'venue', rate_src

    return 0.0, 'unknown_type', rate_src

def was_score(row, tam):
    """Simple WAS scoring mirror — used only to populate the WAS_Scoring_v2 sheet so
    normalize_v3.py has the columns it expects. Real scoring is recomputed when the
    user runs normalize_v3.py over this dataset."""
    typ = row.get('type','')
    score = 3.0  # base
    # Size bump
    if typ == 'Hotel':
        rooms = to_float(row.get('rooms'), 0)
        if rooms >= 400: score += 0.6
        elif rooms >= 250: score += 0.4
        elif rooms >= 150: score += 0.2
    elif typ == 'Restaurant':
        seats = to_float(row.get('seats'), 0)
        if seats >= 300: score += 0.4
        elif seats >= 200: score += 0.2
    elif typ == 'Hospital':
        beds = to_float(row.get('beds'), 0)
        if beds >= 500: score += 0.5
        elif beds >= 250: score += 0.3
    elif typ == 'Venue':
        cap = to_float(row.get('capacity'), 0)
        if cap >= 10000: score += 0.5
        elif cap >= 2000: score += 0.3
    # Operator gate penalty if entrenched national operator
    op = (row.get('valet_operator') or '').lower().strip()
    if op in ('towne park','laz','sp_plus','sp+','propark','park_inc'):
        score -= 0.4
    elif op in ('sophi',):
        score += 1.0
    return round(max(1.0, min(5.0, score)), 2)

# ---------- load ----------
MARKETS = [
    ('ft_wayne', 'Fort Wayne', 'IN'),
    ('cincinnati', 'Cincinnati', 'OH'),
    ('columbus', 'Columbus', 'OH'),
    ('minneapolis', 'Minneapolis', 'MN'),
    ('st_louis', 'St. Louis', 'MO'),
    ('raleigh', 'Raleigh', 'NC'),
    ('charleston', 'Charleston', 'SC'),
    ('ft_lauderdale', 'Fort Lauderdale', 'FL'),
]

all_market_data = {}

# Pass 1: load all rows so we can build the rate pool before computing TAM.
rows_by_market = {}
for key, name, state in MARKETS:
    csv_path = RAW / f'{key}_inventory.csv'
    rows = []
    with open(csv_path) as fp:
        rdr = csv.DictReader(fp)
        for r in rdr:
            r['market_key'] = key
            rows.append(r)
    rows_by_market[key] = rows

RATE_POOL.update(build_rate_pool(rows_by_market))
print(f"Rate pool entries (market,type): {len(RATE_POOL)}")

for key, name, state in MARKETS:
    rows = rows_by_market[key]
    for r in rows:
        tam, basis, rate_src = compute_tam(r)
        r['_tam'] = round(tam, 2)
        r['_tam_basis'] = basis
        r['_rate_source'] = rate_src
        r['_was'] = was_score(r, tam)
    all_market_data[key] = {'name': name, 'state': state, 'rows': rows}

    # ---------- write per-market XLSX ----------
    wb = openpyxl.Workbook()
    # 1) Accounts sheet
    ws = wb.active
    ws.title = 'Accounts'
    accounts_headers = [None,'Account Type','Contact Created?','Address','Uptown (Yes or No)','Phone Number','Email','URL','Self Parking Rate','Valet Rate','# of Rooms','Seats','Beds','Capacity','GM Name','Management Group','Ownership Group','Garage Operator','Valet Operator','Sourcing Notes','Location Notes']
    ws.append(accounts_headers)
    for r in rows:
        ws.append([
            r.get('name'),
            r.get('type'),
            '',
            r.get('address'),
            '',
            r.get('phone'),
            r.get('email'),
            r.get('url'),
            r.get('self_park_rate'),
            r.get('valet_rate'),
            r.get('rooms'),
            r.get('seats'),
            r.get('beds'),
            r.get('capacity'),
            r.get('gm'),
            r.get('management'),
            r.get('ownership_group'),
            r.get('garage_operator'),
            r.get('valet_operator'),
            r.get('sourcing_notes'),
            r.get('location_notes'),
        ])
    # bold header
    for c in ws[1]:
        c.font = Font(bold=True)

    # 2) TAM_Conservative
    ws2 = wb.create_sheet('TAM_Conservative')
    tam_headers = ['Account','Account Type','TAM Class','Rooms/Beds','Seats','Occupancy %','Turnover','Valet Conv %','Valet Rate','Op Days','Peak Mod','TAM','TAM Status','Notes / Source']
    ws2.append(tam_headers)
    for r in rows:
        rooms_or_beds = to_float(r.get('rooms')) or to_float(r.get('beds')) or to_float(r.get('capacity'))
        ws2.append([
            r.get('name'),
            r.get('type'),
            f"{r.get('type')} — {(r.get('location_notes') or '').strip()[:40]}",
            rooms_or_beds,
            to_float(r.get('seats')),
            to_float(r.get('occupancy')),
            to_float(r.get('turnover')),
            to_float(r.get('valet_conv')) or default_valet_conv(r.get('type'), to_float(r.get('rooms'))),
            to_float(r.get('valet_rate')) or to_float(r.get('self_park_rate')),
            to_float(r.get('op_days'), 365),
            to_float(r.get('peak_mod'), 1.0),
            r['_tam'],
            f"Active TAM ({r['_tam_basis']})" if r['_tam'] > 0 else f"No-TAM ({r['_tam_basis']})",
            r.get('sourcing_notes',''),
        ])
    for c in ws2[1]:
        c.font = Font(bold=True)

    # 3) WAS_Scoring_v2
    ws3 = wb.create_sheet('WAS_Scoring_v2')
    was_headers = ['Account','Type','Rooms/Seats/Beds/Cap','Valet Op (incumbent)','Valet Rate','Est. Annual Rev (TAM)','WAS Final','Tier']
    ws3.append(was_headers)
    for r in rows:
        size = to_float(r.get('rooms')) or to_float(r.get('seats')) or to_float(r.get('beds')) or to_float(r.get('capacity'))
        was = r['_was']
        tier = 'A' if was >= 4.0 else ('B' if was >= 3.2 else ('C' if was >= 2.5 else 'D'))
        ws3.append([
            r.get('name'),
            r.get('type'),
            size,
            r.get('valet_operator'),
            to_float(r.get('valet_rate')),
            r['_tam'],
            was,
            tier,
        ])
    for c in ws3[1]:
        c.font = Font(bold=True)

    # widen columns
    for sheet in (ws, ws2, ws3):
        for col in sheet.columns:
            length = max((len(str(c.value)) if c.value else 0) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(40, max(10, length + 2))

    out_path = XLSX_OUT / f'{key}_2026_scored.xlsx'
    wb.save(out_path)

# ---------- combined accounts.json drop-in (normalize_v3 schema) ----------
combined = {'markets': {}, 'meta': {
    'source': 'expansion_v1',
    'generated': '2026-05-07',
    'methodology': 'Charlotte-v3-cookbook TAM formula; comprehensive account universe (hotel+restaurant+venue+hospital).',
}}

for key, name, state in MARKETS:
    md = all_market_data[key]
    accts = []
    tam_total = 0
    for r in md['rows']:
        rec = {
            'name': r.get('name'),
            'market': key,
            'type': r.get('type'),
            'tam': r['_tam'],
            'address': r.get('address'),
            'url': r.get('url'),
            'rooms': to_float(r.get('rooms')),
            'seats': to_float(r.get('seats')),
            'beds': to_float(r.get('beds')),
            'capacity': to_float(r.get('capacity')),
            'self_park_rate': r.get('self_park_rate'),
            'valet_rate': to_float(r.get('valet_rate')),
            'occupancy': to_float(r.get('occupancy')),
            'turnover': to_float(r.get('turnover')),
            'valet_conv': to_float(r.get('valet_conv')) or default_valet_conv(r.get('type'), to_float(r.get('rooms'))),
            'op_days': to_float(r.get('op_days'), 365),
            'peak_mod': to_float(r.get('peak_mod'), 1.0),
            'valet_operator': r.get('valet_operator'),
            'garage_operator': r.get('garage_operator'),
            'management': r.get('management'),
            'ownership_group': r.get('ownership_group'),
            'gm': r.get('gm'),
            'email': r.get('email'),
            'phone': r.get('phone'),
            'location_notes': r.get('location_notes'),
            'sourcing_notes': r.get('sourcing_notes'),
            'was': r['_was'],
            'tam_status': (
                f"Active TAM ({r['_tam_basis']})" if r['_tam'] > 0 else f"No-TAM ({r['_tam_basis']})"
            ),
            'rate_source': r.get('_rate_source'),
        }
        accts.append(rec)
        tam_total += r['_tam']
    combined['markets'][key] = {
        'name': name,
        'state': state,
        'accounts': accts,
        'tam_total': round(tam_total, 2),
        'n_accounts': len(accts),
    }

with open(EXPANSION / 'expansion_accounts.json', 'w') as fp:
    json.dump(combined, fp, indent=2, default=str)

# ---------- summary ----------
print(f"\n{'Market':<18}{'Accts':>6}{'No-TAM':>8}{'TAM Total':>16}")
print('-' * 50)
gtotal = 0
gcount = 0
gno_tam = 0
for key, _, _ in MARKETS:
    md = combined['markets'][key]
    no_tam = sum(1 for a in md['accounts'] if a['tam'] == 0)
    print(f"{key:<18}{md['n_accounts']:>6}{no_tam:>8}{md['tam_total']:>16,.0f}")
    gtotal += md['tam_total']
    gcount += md['n_accounts']
    gno_tam += no_tam
print('-' * 50)
print(f"{'TOTAL':<18}{gcount:>6}{gno_tam:>8}{gtotal:>16,.0f}")
print(f"\nXLSX files: {XLSX_OUT}/")
print(f"Combined JSON: {EXPANSION}/expansion_accounts.json")
